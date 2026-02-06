from db import HEADER_COLUMNS, DETAIL_COLUMNS, get_conn, init_db
import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.utils.datetime import from_excel


ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
BACKEND_DIR = os.path.join(ROOT, "backend")
sys.path.insert(0, BACKEND_DIR)


EXCEL_PATH = os.path.join(ROOT, "data.xlsx")


def maybe_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return None
    return None


def load_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        data.append(dict(zip(header, row)))
    return data


def main():
    if not os.path.exists(EXCEL_PATH):
        raise SystemExit(f"Missing {EXCEL_PATH}")

    init_db()
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    headers = load_sheet(wb["SalesOrderHeader"]
                         ) if "SalesOrderHeader" in wb.sheetnames else []
    details = load_sheet(wb["SalesOrderDetail"]
                         ) if "SalesOrderDetail" in wb.sheetnames else []

    with get_conn() as conn:
        if "--reset" in sys.argv:
            conn.execute("DELETE FROM SalesOrderDetail")
            conn.execute("DELETE FROM SalesOrderHeader")
            conn.execute("DELETE FROM Documents")

        for row in headers:
            row["OrderDate"] = maybe_date(row.get("OrderDate"))
            row["DueDate"] = maybe_date(row.get("DueDate"))
            row["ShipDate"] = maybe_date(row.get("ShipDate"))
            values = {col: row.get(col) for col in HEADER_COLUMNS}
            cols = ", ".join(values.keys())
            placeholders = ", ".join(["?"] * len(values))
            conn.execute(
                f"INSERT OR IGNORE INTO SalesOrderHeader ({cols}) VALUES ({placeholders})",
                list(
                    values.values()),
            )

        for row in details:
            values = {col: row.get(col)
                      for col in DETAIL_COLUMNS if col != "SalesOrderDetailID"}
            cols = ", ".join(values.keys())
            placeholders = ", ".join(["?"] * len(values))
            conn.execute(
                f"INSERT INTO SalesOrderDetail ({cols}) VALUES ({placeholders})",
                list(values.values()),
            )

    print(f"Imported {len(headers)} headers and {len(details)} details.")


if __name__ == "__main__":
    main()
